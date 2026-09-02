import frappe
from frappe.utils.pdf import get_pdf
from whatsapp.api import send_whatsapp
from frappe.utils import get_url
import re


@frappe.whitelist()
def get_po_list():

    data = frappe.get_all(
        "Quotation PO",
        fields=[
            "name",
            "vendor",
            "vendor_delivery_date",
            "jewellery_type"
        ],
        order_by="creation desc"
    )

    for d in data:

        d["total_items"] = frappe.db.count(
            "Quotation PO Item",
            {"po_no": d["name"]}
        )

    return data



@frappe.whitelist()
def get_po_details(po_no):

    # ---------------------------------------------------------
    # GET PO
    # ---------------------------------------------------------

    po = frappe.get_doc(
        "Quotation PO",
        po_no
    )
    
    # Get branch name 
    branch_name = "" 
    if po.branch:
    	branch_name = frappe.db.get_value( "Ornate_Branch_Master", po.branch, "branch_name" ) or ""


    # ---------------------------------------------------------
    # GET PO ITEMS + QUOTATION
    # ---------------------------------------------------------

    items = frappe.db.sql("""
        SELECT
            q.*,
            p.name AS po_item_name,
            p.remark

        FROM `tabQuotation PO Item` p

        INNER JOIN `tabQuotation Upload New` q
            ON q.name = p.item

        WHERE p.po_no = %s

        ORDER BY p.creation ASC

    """, (po_no,), as_dict=True)


    # ---------------------------------------------------------
    # GET DIAMOND DETAILS
    # ---------------------------------------------------------

    quotation_names = [
        item.name
        for item in items
    ]


    diamond_map = {}


    if quotation_names:

        diamond_rows = frappe.db.sql("""
            SELECT
                quotation_number,
                diamond_shape,
                diamond_size,
                diamond_pcs,
                diamond_wt,
                diamond_rate,
                diamond_amount

            FROM `tabQuotation Upload Diamond`

            WHERE quotation_number IN %(quotation_names)s

            ORDER BY
                quotation_number,
                name ASC

        """, {
            "quotation_names": quotation_names
        }, as_dict=True)


        # -----------------------------------------------------
        # GROUP DIAMONDS BY QUOTATION
        # -----------------------------------------------------

        for diamond in diamond_rows:

            quotation_number = str(
                diamond.quotation_number
            )

            if quotation_number not in diamond_map:

                diamond_map[quotation_number] = []


            diamond_map[quotation_number].append({

                "diamond_shape":
                    diamond.diamond_shape or "",

                "diamond_size":
                    diamond.diamond_size or "",

                "diamond_pcs":
                    diamond.diamond_pcs or 0,

                "diamond_wt":
                    diamond.diamond_wt or 0,

                "diamond_rate":
                    diamond.diamond_rate or 0,

                "diamond_amount":
                    diamond.diamond_amount or 0

            })


    # ---------------------------------------------------------
    # ATTACH DIAMOND DETAILS TO EACH ITEM
    # ---------------------------------------------------------

    for item in items:

        item["diamond_details"] = diamond_map.get(
            str(item.name),
            []
        )

    stone_map = {}


    if quotation_names:

        stone_rows = frappe.db.sql("""
            SELECT
                quotation_number,
                stone_pcs,
                stone_wt,
                stone_rate,
                stone_amount

            FROM `tabQuotation Upload Stone`

            WHERE quotation_number IN %(quotation_names)s

            ORDER BY
                quotation_number,
                name ASC

        """, {
            "quotation_names": quotation_names
        }, as_dict=True)


        # -----------------------------------------------------
        # GROUP STONES BY QUOTATION
        # -----------------------------------------------------

        for stone in stone_rows:

            quotation_number = str(
                stone.quotation_number
            )

            if quotation_number not in stone_map:

                stone_map[quotation_number] = []


            stone_map[quotation_number].append({

                "stone_pcs":
                    stone.stone_pcs or 0,

                "stone_wt":
                    stone.stone_wt or 0,

                "stone_rate":
                    stone.stone_rate or 0,

                "stone_amount":
                    stone.stone_amount or 0

            })


    # ---------------------------------------------------------
    # ATTACH STONE DETAILS TO EACH ITEM
    # ---------------------------------------------------------

    for item in items:

        item["stone_details"] = stone_map.get(
            str(item.name),
            []
        )

    # ---------------------------------------------------------
    # RETURN
    # ---------------------------------------------------------

    return {

        "po": po.as_dict(),
		"branch_name": branch_name,
        "items": items

    }



@frappe.whitelist()
def generate_pdf(po_no):

    # ---------------------------------------------------------
    # GET PO DETAILS
    # ---------------------------------------------------------
    po_details = get_po_details(po_no)

    po = frappe._dict(po_details["po"])
    items = po_details["items"]
    branch_name = po_details.get("branch_name", "")

    # ---------------------------------------------------------
    # CALCULATE TOTALS
    # ---------------------------------------------------------
    total_net_wt = 0
    total_dia_wt = 0
    total_stone_wt = 0
    total_qty = 0

    for row in items:

        total_net_wt += float(row.get("net_wt") or 0)

        total_qty += float(row.get("qty") or 1)


        for diamond in row.get("diamond_details", []):

            total_dia_wt += float(
                diamond.get("diamond_wt") or 0
            )
        
        for stone in row.get("stone_details", []):

            total_stone_wt += float(
                stone.get("stone_wt") or 0
            )

    # ---------------------------------------------------------
    # GET USER
    # ---------------------------------------------------------
    user_name = frappe.get_doc(
        "User",
        po.owner
    )

    # ---------------------------------------------------------
    # JINJA CONTEXT
    # ---------------------------------------------------------
    context = {
        "doc": po,
        "items": items,
        "branch_name": branch_name,
        "total_qty": total_qty,
        "total_net_wt": total_net_wt,
        "total_dia_wt": total_dia_wt,
        "total_stone_wt": total_stone_wt,
        "user_name": user_name
    }

    # ---------------------------------------------------------
    # HTML TEMPLATE
    # ---------------------------------------------------------
    html = frappe.render_template(
        """
<style>

@page {
    size: A4;
    margin: 10mm;
}

body {
    font-family: Arial, sans-serif;
    color: #000;
}

table {
    border-collapse: collapse;
    width: 100%;
}

th,
td {
    border: 1px solid #000;
    padding: 5px;
    font-size: 12px;
}

th {
    background: #efefef;
    color: #000;
    font-weight: bold;
}

.header-table td {
    border: none;
    padding: 2px;
    font-size: 13px;
}

.company-title {
    text-align: center;
}

.company-title h2 {
    margin: 2px;
}

.diamond-table td,
.diamond-table th {
    font-size: 11px;
    border: 1px solid #000;
}

.total-table td {
    border: none;
    font-size: 13px;
    font-weight: bold;
}

.terms li {
    margin-bottom: 4px;
    font-size: 10px;
}

.item-image {
    width: 75px;
    height: 75px;
    object-fit: contain;
}

.item-table {
    page-break-inside: auto;
}

.item-table tr {
    page-break-inside: avoid;
}

</style>


<!-- =========================================================
     COMPANY HEADER
========================================================= -->

<div class="company-title">

    <h2>
        <b>SHAH VIRCHAND GOVANJI</b>
    </h2>

    <h2 style="font-size:14px;">
        JEWELLERS PVT. LTD.
    </h2>

    <h2 style="font-size:14px;">
        VALSAD - VAPI - SURAT
    </h2>

    <h2 style="font-size:14px;">
        Phone Number : 02632-229999
    </h2>

    <br>

    <h2 style="font-size:18px;">
        <u>PURCHASE ORDER</u>
    </h2>

</div>


<!-- =========================================================
     PO HEADER
========================================================= -->

<table class="header-table">

    <tr>

        <td width="50%">

            PO No :
            <b>
                {{ doc.name }}
            </b>

        </td>

        <td align="right">

            Vendor :
            <b>
                {{ doc.vendor or "" }}
            </b>

        </td>

    </tr>


    <tr>

        <td>

            PO Date :
            <b>
                {{
                    frappe.utils.format_date(
                        doc.creation,
                        "dd-MM-yyyy"
                    )
                }}
            </b>

        </td>

        <td align="right">

            Delivery Date :
            <b>
                {{
                    frappe.utils.format_date(
                        doc.vendor_delivery_date,
                        "dd-MM-yyyy"
                    )
                    if doc.vendor_delivery_date
                    else ""
                }}
            </b>

        </td>

    </tr>


    <tr>

        <td>


        </td>

        <td align="right">

            Total Qty :
            <b>
                {{ total_qty }}
            </b>

        </td>

    </tr>

</table>


<br>


<!-- =========================================================
     ITEM TABLE
========================================================= -->

<table class="item-table">

    <tr>

        <th width="4%">
            Sr
        </th>

        <th width="12%">
            Image
        </th>

        <th width="18%" align="left">
            Item
        </th>

        <th width="8%" align="left">
            Metal
        </th>

        <th width="6%">
            Qty
        </th>

        <th width="10%" align="right">
            Gross Wt
        </th>

        <th width="10%" align="right">
            Net Wt
        </th>

        <th width="12%" align="right">
            Diamond Wt
        </th>

    </tr>


    {% for row in items %}

    <!-- =====================================================
         ITEM MAIN ROW
    ====================================================== -->

    <tr>

        <td
            rowspan="2"
            align="center"
        >
            {{ loop.index }}
        </td>
<td
    rowspan="2"
    align="center"
    style="vertical-align:middle; text-align:center;"
>
    {% if row.image %}
        <div style="width:100%; text-align:center;">
            <img
                src="{{ row.image }}"
                class="item-image"
                style="
                    display:block;
                    width:75px;
                    height:75px;
                    object-fit:contain;
                    margin:0 auto 5px auto;
                "
            >
        </div>
    {% endif %}

    <div style="
        width:100%;
        text-align:center;
        font-weight:bold;
    ">
        {{ row.vendor_design_number or "" }}
    </div>
</td>
        <td>

            <b>
                {{ row.item or row.name or "" }}
            </b>

        </td>


        <td align="left">

            {{ row.metal or "" }}

        </td>


        <td align="center">

            {{ row.qty or 1 }}

        </td>


        <td align="right">

            {{ row.gr_wt or 0 }}

        </td>


        <td align="right">

            {{ row.net_wt or 0 }}

        </td>


        <td align="right">

            {{
                "%.3f"|format(
                    row.diamond_details
                    |sum(attribute="diamond_wt")
                )
            }}

        </td>

    </tr>


    <!-- =====================================================
         DIAMOND / STONE / REMARK ROW
    ====================================================== -->

    <tr>

        <td
            colspan="6"
            style="padding:0;"
        >

            <!-- =============================================
                 DIAMOND DETAILS
            ============================================== -->

            {% if row.diamond_details %}

            <table
                class="diamond-table"
                width="100%"
            >

                <tr
                    style="
                        background:#f8f8f8;
                        font-weight:bold;
                        text-align:center;
                    "
                >

                    <th width="18%">
                        Diamond
                    </th>

                    <th>
                        Shape
                    </th>

                    <th>
                        Size
                    </th>

                    <th>
                        Pcs
                    </th>

                    <th>
                        Weight
                    </th>

                    <th>
                        Rate
                    </th>

                    <th>
                        Amount
                    </th>

                </tr>


                {% for diamond in row.diamond_details %}

                <tr align="center">

                    <td>

                        <b>
                            Diamond {{ loop.index }}
                        </b>

                    </td>

                    <td>
                        {{ diamond.diamond_shape or "" }}
                    </td>

                    <td>
                        {{ diamond.diamond_size or "" }}
                    </td>

                    <td>
                        {{ diamond.diamond_pcs or "" }}
                    </td>

                    <td>
                        {{ diamond.diamond_wt or "" }}
                    </td>

                    <td>
                        {{ diamond.diamond_rate or "" }}
                    </td>

                    <td>
                        {{ diamond.diamond_amount or "" }}
                    </td>

                </tr>

                {% endfor %}

            </table>

            {% endif %}


            <!-- =============================================
                 STONE DETAILS
            ============================================== -->

            {% if row.stone_details %}

            <table
                class="stone-table"
                width="100%"
            >

                <tr
                    style="
                        background:#f8f8f8;
                        font-weight:bold;
                        text-align:center;
                    "
                >

                    <th width="18%">
                        Stone
                    </th>

                    <th>
                        Pcs
                    </th>

                    <th>
                        Weight
                    </th>

                    <th>
                        Rate
                    </th>

                    <th>
                        Amount
                    </th>

                </tr>


                {% for stone in row.stone_details %}

                <tr align="center">

                    <td>

                        <b>
                            Diamond {{ loop.index }}
                        </b>

                    </td>

                    <td>
                        {{ stone.stone_pcs or "" }}
                    </td>

                    <td>
                        {{ stone.stone_wt or "" }}
                    </td>

                    <td>
                        {{ stone.stone_rate or "" }}
                    </td>

                    <td>
                        {{ stone.stone_amount or "" }}
                    </td>

                </tr>

                {% endfor %}

            </table>

            {% endif %}



            <!-- =============================================
                 REMARK
            ============================================== -->

            <table
                class="diamond-table"
                width="100%"
            >

                <tr>

                    <td
                        width="18%"
                        style="font-weight:bold;"
                    >
                        Remark
                    </td>

                    <td colspan="6">

                        {{ row.remark or "" }}

                    </td>

                </tr>

            </table>

        </td>

    </tr>


    {% endfor %}

</table>


<br>


<!-- =========================================================
     TOTALS
========================================================= -->

<table class="total-table">

    <tr>

        <td align="right">

            Total Qty :
            {{ total_qty }}

            &nbsp;&nbsp;&nbsp;&nbsp;

            Total Net Wt :
            {{ "%.3f"|format(total_net_wt) }}

            &nbsp;&nbsp;&nbsp;&nbsp;

            Total Diamond Wt :
            {{ "%.3f"|format(total_dia_wt) }}

            &nbsp;&nbsp;&nbsp;&nbsp;

            Total Stone Wt :
            {{ "%.3f"|format(total_stone_wt) }}

        </td>

    </tr>

</table>


<br>


<!-- =========================================================
     TERMS
========================================================= -->

<div style="font-weight:bold;">
    Important Terms and Conditions
</div>


<ul class="terms">

    <li>
        <b>
        The weight mentioned in the Purchase Order is final,
        and the order must be as per the specified weight.
        </b>
    </li>

    <li>
        <b>
        Delivery dates specified in the PO are binding.
        </b>
    </li>

    <li>
        <b>
        Goods must conform to agreed specifications and be free
        from defects.
        </b>
    </li>

    <li>
        <b>
        Rejected goods will be returned at the vendor's expense.
        </b>
    </li>

    <li>
        <b>
        The vendor is responsible for all expenses related to
        repairing jewellery products.
        </b>
    </li>

    <li>
        <b>
        The buyer reserves the right to cancel the PO without
        liability if the vendor fails to meet agreed-upon terms
        or specifications.
        </b>
    </li>

    <li>
        <b>
        Please review the Purchase Order and confirm your
        acceptance by replying to this email or via WhatsApp
        at your earliest convenience.
        </b>
    </li>

    <li>
        <b>
        If a product is damaged or has a QC issue and VG repairs it
        at its facility, charges for the same will be borne by the
        vendor.
        </b>
    </li>

</ul>


<br><br><br>


<!-- =========================================================
     SIGNATURE
========================================================= -->

<table class="header-table">

    <tr>

        <td width="70%"></td>

        <td align="right">

            <u>
                {{ user_name.full_name }}
            </u>

        </td>

    </tr>


    <tr>

        <td></td>

        <td align="right">

            (Purchase Department)

        </td>

    </tr>

</table>
        """,
        context
    )

    # ---------------------------------------------------------
    # GENERATE PDF
    # ---------------------------------------------------------
    pdf = get_pdf(html)

    # ---------------------------------------------------------
    # SAVE PDF
    # ---------------------------------------------------------

    safe_po_no = str(po_no).replace("/", "-")

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": f"{safe_po_no}.pdf",
        "attached_to_doctype": "Quotation PO",
        "attached_to_name": po_no,
        "is_private": 0,
        "content": pdf
    })

    file_doc.save(ignore_permissions=True)

    frappe.db.commit()

    return {
        "success": True,
        "file_url": file_doc.file_url
    }




@frappe.whitelist(allow_guest=True)    
def send_pending_po_emails():
    po_list = frappe.get_all(
        "Quotation PO",
        filters={
            "is_mail_send": 1,
            "status": 1
        },
        fields=[
            "name",
            "po_no",
            "vendor",
            "vendor_code",
            "vendor_delivery_date",
            "branch"
        ],
        order_by="creation asc"
    )

    if not po_list:
        return

    for po in po_list:

        try:

            # -------------------------------------------------
            # GET FULL PO
            # -------------------------------------------------
            po_doc = frappe.get_doc(
                "Quotation PO",
                po.name
            )



            vendor_code = po_doc.vendor_code
            po_format ="Pdf"

            if vendor_code:
                vendor_code = re.sub(
                    r'^(\d{2})-',
                    r'0\1-',
                    str(vendor_code)
                )
            # -------------------------------------------------
            # GET VENDOR EMAIL
            #
            # PO.vendor = supplier_code
            # -------------------------------------------------

            vendor_data = frappe.db.get_value(
                "Ornate_Supplier_Master",
                {
                    "supplier_code": vendor_code
                },
                ["po_contact_person_email", "supplier_name","po_contact_person_mobile1","po_format"],
                as_dict=True
            )

            vendor_email = vendor_data.po_contact_person_email if vendor_data else None
            vendor_name = vendor_data.supplier_name if vendor_data else None
            vendor_mobile = vendor_data.po_contact_person_mobile1 if vendor_data else None
            po_format = vendor_data.po_format if vendor_data else "Pdf"


            user_name = frappe.get_doc( "User", po_doc.owner )


            # Fallback to Visitor Vendor
            if not vendor_email or not vendor_name:
                visitor_vendor = frappe.db.get_value(
                    "Visitor Vendor",
                    {
                        "vendor_code": po_doc.vendor_code
                    },
                    ["vendor_email", "vendor_name","vendor_mobile_no"],
                    as_dict=True
                )

                if visitor_vendor:
                    if not vendor_email:
                        vendor_email = visitor_vendor.vendor_email

                    if not vendor_name:
                        vendor_name = visitor_vendor.vendor_name
                    
                    if not vendor_mobile:
                        vendor_mobile = visitor_vendor.vendor_mobile_no
            
            is_whatsapp_send= False
            if not vendor_email:
                vendor_param= [po.name,user_name.full_name,po_doc.vendor_code]
                send_whatsapp("919512152521","po_email_failed_vendor",vendor_param)
                send_whatsapp("919273446652","po_email_failed_vendor",vendor_param)
                send_whatsapp("918238095376","po_email_failed_vendor",vendor_param)
                is_whatsapp_send = True
                frappe.log_error(
                    f"Vendor email not found. "
                    f"Vendor Code: {po_doc.vendor}, "
                    f"PO: {po.name}",
                    "PO Email Cron"
                )

            
            if not vendor_mobile:
                if not is_whatsapp_send:
                    vendor_param= [po.name,user_name.full_name,po_doc.vendor_code]
                    send_whatsapp("919512152521","po_email_failed_vendor",vendor_param)
                    send_whatsapp("919273446652","po_email_failed_vendor",vendor_param)
                    send_whatsapp("918238095376","po_email_failed_vendor",vendor_param)
                frappe.log_error(
                    f"Vendor mobile not found. "
                    f"Vendor Code: {po_doc.vendor}, "
                    f"PO: {po.name}",
                    "PO Mobile Cron"
                )

            if not vendor_email or not vendor_mobile:
                continue
            # -------------------------------------------------
            # GENERATE PDF
            # -------------------------------------------------
            if po_format =="Excel":
                pdf_response = generate_excel(po.name)
            else:
                pdf_response = generate_pdf(po.name)
            if not pdf_response:
                continue

            file_url = pdf_response.get("file_url")

            if not file_url:
                continue
            base_url = frappe.utils.get_url()
            pdf_url = f"{base_url}{file_url}"
            link = " "
            body_param =[vendor_name,po.name,"8238095376",link ]
            mobile = f"91{vendor_mobile}"
            send_whatsapp(mobile,"purchase_order_whatsapp_with_link_new",pdf_url,body_param)
            send_whatsapp("918238095376","purchase_order_whatsapp_with_link_new",pdf_url,body_param)
            send_whatsapp("919273446652","purchase_order_whatsapp_with_link_new",pdf_url,body_param)
            send_whatsapp("919512152521","purchase_order_whatsapp_with_link_new",pdf_url,body_param)


            # -------------------------------------------------
            # GET PDF FILE
            # -------------------------------------------------
            file_doc = frappe.get_doc(
                "File",
                {
                    "file_url": file_url
                }
            )

            # -------------------------------------------------
            # EMAIL
            # -------------------------------------------------
            subject = f"Purchase Order - {po.name}"

            message = f"""
                <p>Dear Sir/Madam,</p>

                <p>
                    Please find attached Purchase Order
                    <b>{po.name}</b>.
                </p>

                <p>
                    Kindly review the Purchase Order and confirm
                    your acceptance.
                </p>

                <br>

                <p>
                    Regards,<br>
                    Purchase Department<br>
                    SHAH VIRCHAND GOVANJI JEWELLERS PVT. LTD.
                </p>
            """

            # -------------------------------------------------
            # SEND EMAIL
            # -------------------------------------------------
            frappe.sendmail(
                recipients=[vendor_email,"itdigital@svgjewels.com"],
                cc =["mukesh.k@svgjewels.com","miteshthakur87@gmail.com","diamond@svgjewels.com"],
                reply_to="diamond@svgjewels.com",
                subject=subject,
                message=message,
                attachments=[
                    {
                        "fname": file_doc.file_name,
                        "fcontent": file_doc.get_content()
                    }
                ],
                reference_doctype="Quotation PO",
                reference_name=po.name
            )

            # -------------------------------------------------
            # MARK AS MAIL SENT
            # -------------------------------------------------
            frappe.db.set_value(
                "Quotation PO",
                po.name,
                "is_mail_send",
                0
            )

            frappe.db.commit()

        except Exception:

            frappe.log_error(
                frappe.get_traceback(),
                f"PO Email Failed - {po.name}"
            )

            frappe.db.rollback()    

@frappe.whitelist(allow_guest=True)    
def send_pending_whatsapp():
    po_list = frappe.get_all(
        "Quotation PO",
        filters={
            "is_mail_send": 0,
            "status": 1,

        },
        fields=[
            "name",
            "po_no",
            "vendor",
            "vendor_code",
            "vendor_delivery_date",
            "branch"
        ],
        order_by="creation asc"
    )

    if not po_list:
        return

    for po in po_list:

        try:

            # -------------------------------------------------
            # GET FULL PO
            # -------------------------------------------------
            po_doc = frappe.get_doc(
                "Quotation PO",
                po.name
            )


            po_format = "Pdf" 
            vendor_code = po_doc.vendor_code

            if(vendor_code == "093-DIAMOND" ):
                continue

            if vendor_code:
                vendor_code = re.sub(
                    r'^(\d{2})-',
                    r'0\1-',
                    str(vendor_code)
                )
            # -------------------------------------------------
            # GET VENDOR EMAIL
            #
            # PO.vendor = supplier_code
            # -------------------------------------------------

            vendor_data = frappe.db.get_value(
                "Ornate_Supplier_Master",
                {
                    "supplier_code": vendor_code
                },
                ["po_contact_person_email", "supplier_name","po_contact_person_mobile1","po_format"],
                as_dict=True
            )

            vendor_email = vendor_data.po_contact_person_email if vendor_data else None
            vendor_name = vendor_data.supplier_name if vendor_data else None
            vendor_mobile = vendor_data.po_contact_person_mobile1 if vendor_data else None
            po_format = vendor_data.po_format if vendor_data else "Pdf"


            # Fallback to Visitor Vendor
            if not vendor_email or not vendor_name:
                visitor_vendor = frappe.db.get_value(
                    "Visitor Vendor",
                    {
                        "vendor_code": po_doc.vendor_code
                    },
                    ["vendor_email", "vendor_name","vendor_mobile_no"],
                    as_dict=True
                )

                if visitor_vendor:
                    if not vendor_email:
                        vendor_email = visitor_vendor.vendor_email

                    if not vendor_name:
                        vendor_name = visitor_vendor.vendor_name
                    
                    if not vendor_mobile:
                        vendor_mobile = visitor_vendor.vendor_mobile_no
            
            if not vendor_email:

                frappe.log_error(
                    f"Vendor email not found. "
                    f"Vendor Code: {po_doc.vendor}, "
                    f"PO: {po.name}",
                    "PO Email Cron"
                )

                continue

            # -------------------------------------------------
            # GENERATE PDF
            # -------------------------------------------------

            if po_format =="Excel":
                pdf_response = generate_excel(po.name)
            else:
                pdf_response = generate_pdf(po.name)
            if not pdf_response:
                continue

            file_url = pdf_response.get("file_url")

            if not file_url:
                continue
            base_url = frappe.utils.get_url()
            pdf_url = f"{base_url}{file_url}"
            link = " "
            body_param =[vendor_name,po.name,"8238095376",link ]
            mobile = f"91{vendor_mobile}"
            send_whatsapp(mobile,"purchase_order_whatsapp_with_link_new",pdf_url,body_param)
            send_whatsapp("919273446652","purchase_order_whatsapp_with_link_new",pdf_url,body_param)
            send_whatsapp("919512152521","purchase_order_whatsapp_with_link_new",pdf_url,body_param)


        except Exception:

            frappe.log_error(
                frappe.get_traceback(),
                f"PO Email Failed - {po.name}"
            )

            frappe.db.rollback()    

@frappe.whitelist()
def generate_excel(po_no):

    import io
    import os
    import requests

    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        Alignment,
        Border,
        Side,
        PatternFill
    )
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    # ---------------------------------------------------------
    # GET PO DETAILS
    # ---------------------------------------------------------

    po_details = get_po_details(po_no)

    po = frappe._dict(po_details["po"])
    items = po_details["items"]
    branch_name = po_details.get("branch_name", "")

    # ---------------------------------------------------------
    # CALCULATE TOTALS
    # ---------------------------------------------------------

    total_net_wt = 0
    total_dia_wt = 0
    total_stone_wt = 0
    total_qty = 0

    for row in items:

        total_net_wt += float(
            row.get("net_wt") or 0
        )

        total_qty += float(
            row.get("qty") or 1
        )

        for diamond in row.get("diamond_details", []):

            total_dia_wt += float(
                diamond.get("diamond_wt") or 0
            )

        for stone in row.get("stone_details", []):

            total_stone_wt += float(
                stone.get("stone_wt") or 0
            )

    # ---------------------------------------------------------
    # GET USER
    # ---------------------------------------------------------

    user_name = frappe.get_doc(
        "User",
        po.owner
    )

    # ---------------------------------------------------------
    # CREATE WORKBOOK
    # ---------------------------------------------------------

    wb = Workbook()

    ws = wb.active
    ws.title = "Purchase Order"

    # ---------------------------------------------------------
    # PAGE SETTINGS
    # ---------------------------------------------------------

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    # ---------------------------------------------------------
    # COLUMN WIDTHS
    # ---------------------------------------------------------

    widths = {
        "A": 6,     # Sr
        "B": 18,    # Image
        "C": 22,    # Item
        "D": 16,    # Metal
        "E": 8,     # Qty
        "F": 13,    # Gross
        "G": 13,    # Net
        "H": 15,    # Diamond
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ---------------------------------------------------------
    # STYLES
    # ---------------------------------------------------------

    thin = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    no_border = Border()

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="EFEFEF"
    )

    detail_fill = PatternFill(
        fill_type="solid",
        fgColor="F8F8F8"
    )

    title_font = Font(
        name="Arial",
        size=16,
        bold=True
    )

    subtitle_font = Font(
        name="Arial",
        size=12,
        bold=True
    )

    normal_font = Font(
        name="Arial",
        size=10
    )

    bold_font = Font(
        name="Arial",
        size=10,
        bold=True
    )

    small_font = Font(
        name="Arial",
        size=9
    )

    # ---------------------------------------------------------
    # COMPANY HEADER
    # ---------------------------------------------------------

    ws.merge_cells("A1:H1")

    ws["A1"] = "SHAH VIRCHAND GOVANJI"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("A2:H2")

    ws["A2"] = "JEWELLERS PVT. LTD."
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("A3:H3")

    ws["A3"] = "VALSAD - VAPI - SURAT"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("A4:H4")

    ws["A4"] = "Phone Number : 02632-229999"
    ws["A4"].font = subtitle_font
    ws["A4"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("A6:H6")

    ws["A6"] = "PURCHASE ORDER"
    ws["A6"].font = Font(
        name="Arial",
        size=15,
        bold=True,
        underline="single"
    )

    ws["A6"].alignment = Alignment(
        horizontal="center"
    )

    # ---------------------------------------------------------
    # PO HEADER
    # ---------------------------------------------------------

    ws["A8"] = "PO No :"
    ws["A8"].font = bold_font

    ws["B8"] = po.name
    ws["B8"].font = bold_font

    ws["E8"] = "Vendor :"
    ws["E8"].font = bold_font

    ws["H8"] = po.vendor or ""
    ws["H8"].font = bold_font
    ws["H8"].alignment = Alignment(
        horizontal="right"
    )

    ws["A9"] = "PO Date :"
    ws["A9"].font = bold_font

    ws["B9"] = (
        frappe.utils.format_date(
            po.creation,
            "dd-MM-yyyy"
        )
    )

    ws["E9"] = "Delivery Date :"
    ws["E9"].font = bold_font

    ws["H9"] = (
        frappe.utils.format_date(
            po.vendor_delivery_date,
            "dd-MM-yyyy"
        )
        if po.vendor_delivery_date
        else ""
    )

    ws["A10"] = ""
    ws["E10"] = "Total Qty :"
    ws["E10"].font = bold_font

    ws["H10"] = total_qty
    ws["H10"].font = bold_font

    # ---------------------------------------------------------
    # ITEM TABLE HEADER
    # ---------------------------------------------------------

    start_row = 12

    headers = [
        "Sr",
        "Image",
        "Item",
        "Metal",
        "Qty",
        "Gross Wt",
        "Net Wt",
        "Diamond Wt"
    ]

    for col_num, value in enumerate(headers, 1):

        cell = ws.cell(
            row=start_row,
            column=col_num,
            value=value
        )

        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[start_row].height = 25

    current_row = start_row + 1

    # ---------------------------------------------------------
    # ITEMS
    # ---------------------------------------------------------

    for index, row in enumerate(items, 1):

        item_start_row = current_row
        item_end_row = current_row + 1

        # -----------------------------------------------------
        # MAIN ROW
        # -----------------------------------------------------

        values = [
            index,
            "",
            row.get("item") or row.get("name") or "",
            row.get("metal") or "",
            row.get("qty") or 1,
            float(row.get("gr_wt") or 0),
            float(row.get("net_wt") or 0),
            sum(
                float(
                    d.get("diamond_wt") or 0
                )
                for d in row.get(
                    "diamond_details",
                    []
                )
            )
        ]

        for col_num, value in enumerate(values, 1):

            cell = ws.cell(
                row=item_start_row,
                column=col_num,
                value=value
            )

            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # Item should be left aligned
        ws.cell(
            item_start_row,
            3
        ).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        # ---------------------------------------------------------
        # IMAGE
        # ---------------------------------------------------------

        image_path = row.get("image")

        if image_path:

            try:

                image_data = None

                # =================================================
                # CASE 1: /files/xxxx.jpg
                # =================================================

                if image_path.startswith("/files/"):

                    local_path = frappe.get_site_path(
                        "public",
                        image_path.lstrip("/")
                    )

                    if os.path.exists(local_path):

                        with open(
                            local_path,
                            "rb"
                        ) as f:

                            image_data = f.read()


                # =================================================
                # CASE 2: /private/files/xxxx.jpg
                # =================================================

                elif image_path.startswith("/private/files/"):

                    local_path = frappe.get_site_path(
                        image_path.lstrip("/")
                    )


                    if os.path.exists(local_path):

                        with open(
                            local_path,
                            "rb"
                        ) as f:

                            image_data = f.read()


                # =================================================
                # CASE 3: HTTP / HTTPS
                # =================================================

                elif (
                    image_path.startswith("http://")
                    or
                    image_path.startswith("https://")
                ):

                    response = requests.get(
                        image_path,
                        timeout=30
                    )

                    response.raise_for_status()

                    image_data = response.content


                # =================================================
                # CASE 4: LOCAL FILE PATH
                # =================================================

                elif os.path.exists(image_path):

                    with open(
                        image_path,
                        "rb"
                    ) as f:

                        image_data = f.read()


                # =================================================
                # ADD IMAGE TO EXCEL
                # =================================================

                if image_data:

                    image_stream = io.BytesIO(
                        image_data
                    )

                    xl_image = XLImage(
                        image_stream
                    )

                    # Excel image size
                    xl_image.width = 75
                    xl_image.height = 75

                    # Add image
                    ws.add_image(
                        xl_image,
                        f"B{item_start_row}"
                    )

                else:

                    frappe.log_error(
                        f"Image not found: {image_path}",
                        "PO Excel Image"
                    )


            except Exception:

                frappe.log_error(
                    frappe.get_traceback(),
                    "PO Excel Image Error"
                )
        # -----------------------------------------------------
        # VENDOR DESIGN NUMBER
        # -----------------------------------------------------

        design_no = (
            row.get("vendor_design_number")
            or ""
        )

        ws.cell(
            item_start_row,
            2
        ).value = design_no

        ws.cell(
            item_start_row,
            2
        ).font = bold_font

        ws.cell(
            item_start_row,
            2
        ).alignment = Alignment(
            horizontal="center",
            vertical="bottom"
        )

        # -----------------------------------------------------
        # SECOND ROW
        # -----------------------------------------------------

        for col_num in range(1, 9):

            cell = ws.cell(
                item_end_row,
                col_num
            )

            cell.border = border
            cell.fill = detail_fill
            cell.font = small_font
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

        # -----------------------------------------------------
        # DIAMOND DETAILS
        # -----------------------------------------------------

        detail_row = item_end_row

        diamond_details = row.get(
            "diamond_details",
            []
        )

        if diamond_details:

            ws.cell(
                detail_row,
                3,
                "DIAMOND DETAILS"
            )

            ws.cell(
                detail_row,
                3
            ).font = bold_font

            diamond_header_row = detail_row + 1

            diamond_headers = [
                "Diamond",
                "Shape",
                "Size",
                "Pcs",
                "Weight",
                "Rate",
                "Amount"
            ]

            # Put diamond table in C:H
            diamond_start_col = 3

            for i, header in enumerate(
                diamond_headers
            ):

                col = diamond_start_col + i

                if col > 8:
                    break

                cell = ws.cell(
                    diamond_header_row,
                    col,
                    header
                )

                cell.font = bold_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            detail_row = diamond_header_row + 1

            for dia_index, diamond in enumerate(
                diamond_details,
                1
            ):

                diamond_values = [
                    f"Diamond {dia_index}",
                    diamond.get(
                        "diamond_shape"
                    ) or "",
                    diamond.get(
                        "diamond_size"
                    ) or "",
                    diamond.get(
                        "diamond_pcs"
                    ) or 0,
                    float(
                        diamond.get(
                            "diamond_wt"
                        ) or 0
                    ),
                    float(
                        diamond.get(
                            "diamond_rate"
                        ) or 0
                    ),
                    float(
                        diamond.get(
                            "diamond_amount"
                        ) or 0
                    )
                ]

                for i, value in enumerate(
                    diamond_values
                ):

                    col = diamond_start_col + i

                    if col > 8:
                        break

                    cell = ws.cell(
                        detail_row,
                        col,
                        value
                    )

                    cell.border = border
                    cell.font = small_font
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

                detail_row += 1

        # -----------------------------------------------------
        # STONE DETAILS
        # -----------------------------------------------------

        if row.get("stone_details"):

            detail_row += 1

            ws.cell(
                detail_row,
                3,
                "STONE DETAILS"
            )

            ws.cell(
                detail_row,
                3
            ).font = bold_font

            stone_header_row = detail_row + 1

            stone_headers = [
                "Stone",
                "Pcs",
                "Weight",
                "Rate",
                "Amount"
            ]

            for i, header in enumerate(
                stone_headers
            ):

                col = 3 + i

                cell = ws.cell(
                    stone_header_row,
                    col,
                    header
                )

                cell.font = bold_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center"
                )

            detail_row = stone_header_row + 1

            for stone_index, stone in enumerate(
                row.get("stone_details", []),
                1
            ):

                stone_values = [
                    f"Stone {stone_index}",
                    stone.get(
                        "stone_pcs"
                    ) or 0,
                    float(
                        stone.get(
                            "stone_wt"
                        ) or 0
                    ),
                    float(
                        stone.get(
                            "stone_rate"
                        ) or 0
                    ),
                    float(
                        stone.get(
                            "stone_amount"
                        ) or 0
                    )
                ]

                for i, value in enumerate(
                    stone_values
                ):

                    col = 3 + i

                    cell = ws.cell(
                        detail_row,
                        col,
                        value
                    )

                    cell.border = border
                    cell.font = small_font
                    cell.alignment = Alignment(
                        horizontal="center"
                    )

                detail_row += 1

        # -----------------------------------------------------
        # REMARK
        # -----------------------------------------------------

        detail_row += 1

        ws.cell(
            detail_row,
            3,
            "Remark"
        )

        ws.cell(
            detail_row,
            3
        ).font = bold_font

        ws.merge_cells(
            start_row=detail_row,
            start_column=4,
            end_row=detail_row,
            end_column=8
        )

        remark_cell = ws.cell(
            detail_row,
            4
        )

        remark_cell.value = (
            row.get("remark") or ""
        )

        remark_cell.font = normal_font

        remark_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

        # Add border around remark
        for col in range(3, 9):

            ws.cell(
                detail_row,
                col
            ).border = border

        # -----------------------------------------------------
        # ROW HEIGHT
        # -----------------------------------------------------

        ws.row_dimensions[
            item_start_row
        ].height = 90

        # Move to next item
        current_row = detail_row + 2

    # ---------------------------------------------------------
    # TOTALS
    # ---------------------------------------------------------

    total_row = current_row

    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=8
    )

    total_cell = ws.cell(
        total_row,
        1
    )

    total_cell.value = (
        f"Total Qty : {total_qty}    "
        f"Total Net Wt : {total_net_wt:.3f}    "
        f"Total Diamond Wt : {total_dia_wt:.3f}    "
        f"Total Stone Wt : {total_stone_wt:.3f}"
    )

    total_cell.font = Font(
        name="Arial",
        size=11,
        bold=True
    )

    total_cell.alignment = Alignment(
        horizontal="right"
    )

    # ---------------------------------------------------------
    # TERMS
    # ---------------------------------------------------------

    terms_start = total_row + 3

    ws.merge_cells(
        start_row=terms_start,
        start_column=1,
        end_row=terms_start,
        end_column=8
    )

    ws.cell(
        terms_start,
        1,
        "Important Terms and Conditions"
    )

    ws.cell(
        terms_start,
        1
    ).font = Font(
        name="Arial",
        size=11,
        bold=True
    )

    terms = [
        "The weight mentioned in the Purchase Order is final, and the order must be as per the specified weight.",
        "Delivery dates specified in the PO are binding.",
        "Goods must conform to agreed specifications and be free from defects.",
        "Rejected goods will be returned at the vendor's expense.",
        "The vendor is responsible for all expenses related to repairing jewellery products.",
        "The buyer reserves the right to cancel the PO without liability if the vendor fails to meet agreed-upon terms or specifications.",
        "Please review the Purchase Order and confirm your acceptance by replying to this email or via WhatsApp at your earliest convenience.",
        "If a product is damaged or has a QC issue and VG repairs it at its facility, charges for the same will be borne by the vendor."
    ]

    term_row = terms_start + 1

    for index, term in enumerate(
        terms,
        1
    ):

        ws.merge_cells(
            start_row=term_row,
            start_column=1,
            end_row=term_row,
            end_column=8
        )

        cell = ws.cell(
            term_row,
            1
        )

        cell.value = f"{index}. {term}"

        cell.font = Font(
            name="Arial",
            size=9,
            bold=True
        )

        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

        term_row += 1

    # ---------------------------------------------------------
    # SIGNATURE
    # ---------------------------------------------------------

    signature_row = term_row + 3

    ws.merge_cells(
        start_row=signature_row,
        start_column=6,
        end_row=signature_row,
        end_column=8
    )

    signature_cell = ws.cell(
        signature_row,
        6
    )

    signature_cell.value = (
        user_name.full_name
    )

    signature_cell.font = Font(
        name="Arial",
        size=10,
        bold=False,
        underline="single"
    )

    signature_cell.alignment = Alignment(
        horizontal="right"
    )

    ws.merge_cells(
        start_row=signature_row + 1,
        start_column=6,
        end_row=signature_row + 1,
        end_column=8
    )

    ws.cell(
        signature_row + 1,
        6
    ).value = "(Purchase Department)"

    ws.cell(
        signature_row + 1,
        6
    ).font = Font(
        name="Arial",
        size=10
    )

    ws.cell(
        signature_row + 1,
        6
    ).alignment = Alignment(
        horizontal="right"
    )

    # ---------------------------------------------------------
    # NUMBER FORMATS
    # ---------------------------------------------------------

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(
                cell.value,
                float
            ):

                cell.number_format = "0.000"

    # ---------------------------------------------------------
    # FREEZE
    # ---------------------------------------------------------

    #ws.freeze_panes = "A13"

    # ---------------------------------------------------------
    # SAVE TO MEMORY
    # ---------------------------------------------------------

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    excel_content = output.getvalue()

    # ---------------------------------------------------------
    # SAFE FILE NAME
    # ---------------------------------------------------------

    safe_po_no = str(
        po_no
    ).replace(
        "/",
        "-"
    )

    file_name = (
        f"{safe_po_no}.xlsx"
    )

    # ---------------------------------------------------------
    # SAVE FILE IN FRAPPE
    # ---------------------------------------------------------

    # Delete existing file with same name and attachment
    """existing_file = frappe.db.exists("File", {
        "file_name": file_name,
        "attached_to_doctype": "Quotation PO",
        "attached_to_name": po_no
    })

    if existing_file:
        frappe.delete_doc(
            "File",
            existing_file,
            ignore_permissions=True,
            force=True
        )"""

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "attached_to_doctype": "Quotation PO",
        "attached_to_name": po_no,
        "is_private": 0,
        "content": excel_content
    })

    file_doc.save(
        ignore_permissions=True
    )

    frappe.db.commit()

    # ---------------------------------------------------------
    # RETURN
    # ---------------------------------------------------------

    return {
        "success": True,
        "file_url": file_doc.file_url,
        "file_name": file_name
    }
